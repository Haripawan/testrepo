from openpyxl import load_workbook
from openpyxl.styles import Font

def highlight_non_ascii(file_path, sheet_name=None):
    # Load the workbook and specify the sheet
    workbook = load_workbook(filename=file_path)
    sheet = workbook[sheet_name] if sheet_name else workbook.active
    
    # Define a red font for highlighting
    red_font = Font(color="FF0000")
    
    # Iterate over each cell in the sheet
    for row in sheet.iter_rows():
        for cell in row:
            # Check if the cell contains non-ASCII characters
            if isinstance(cell.value, str) and any(ord(char) > 127 for char in cell.value):
                cell.font = red_font
    
    # Save the workbook
    workbook.save(file_path)

# Example usage
highlight_non_ascii("your_excel_file.xlsx", "Sheet1")